from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


mes = "Junho" # variavel com o mes que vamos analisar
wb = load_workbook("tabela_produtos.xlsx") # carrega a biblioteca no arquivo excel
sheet = wb["Report"]  # seleciona qual sheet do excel vc quer utilizar


min_column = wb.active.min_column  # cria variavel pra cada valor das colunas e linhas
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()  # cria variavel com a função da barchart

data = Reference(sheet,          # cria a variavel com as referencias da posição dos dados
          min_col=min_column+1,  # selecionando a segunda menor coluna para separar os dois generos e não a celula escrita genero
          max_col=max_column,
          min_row=min_row,
          max_row=max_row)

categorias = Reference(sheet,
          min_col=min_column,
          max_col=min_column,
          min_row=min_row+1, # selecionando a segunda linhas para pegar as categorias de cada um
          max_row=max_row)

barchart.add_data(data, titles_from_data=True)  # colocando os dados dentro da barchart
barchart.set_categories(categorias)  # colocando as categorias que os dados serão inseridos dentro

sheet.add_chart(barchart, "B12") # criando a barchart com a linha e selecionando a linha em que ela vai começar


barchart.title = "Sales by Product line"  # o nome da barchart
barchart.style = 2  # o estilo que gostaria


for i in range(min_column+1, max_column+1): # criando um for loop onde ele vai pegar a coluna minima mais 1 e a coluna maxima
    #  mexendo com range a coluna maxima sempre é ignorada então coloque a maxima +1
    letter = get_column_letter(i)  # criando a variavel para guardar a letra da coluna
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    # botando dentro do sheet a letra e a linha onde serão colocados os dados
    sheet[f"{letter}{max_row+1}"].style = "Currency"


sheet["A1"] = "Relatorio de Vendas"
sheet["A1"].font = Font("Arial", bold=True, size=20)
sheet["A2"] = mes
sheet["A2"].font = Font("Arial", bold=True, size=10)

wb.save(f"report_{mes}.xlsx")